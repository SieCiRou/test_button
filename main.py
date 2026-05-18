import os
import time
import ctypes
import numpy as np
from pywinauto.application import Application
import win32gui
import win32api
from openpyxl import load_workbook, Workbook  # 引入 Excel 處理庫

# 確保高 DPI 螢幕真實實體座標精確
ctypes.windll.shcore.SetProcessDpiAwareness(2)

def get_pixel_color(hdc, x, y):
    """使用 Windows GDI API 快速獲取指定座標的 RGB 顏色"""
    color = win32gui.GetPixel(hdc, x, y)
    b = (color >> 16) & 0xff
    g = (color >> 8) & 0xff
    r = color & 0xff
    return (r, g, b)

def is_slate_gray(rgb):
    """判斷是否為未選中狀態的 SlateGray (特色：R, G, B 各值比較接近，偏灰暗)"""
    r, g, b = rgb
    return (90 <= r <= 130) and (110 <= g <= 150) and (130 <= b <= 165)

def is_steel_blue(rgb):
    """判斷是否為真正有作用的 SteelBlue (特色：R值顯著下降，且B值比灰色更飽和)
    此範圍已排除 LightSteelBlue (LightSteelBlue 的 R 通常大於 160，顏色太亮)"""
    r, g, b = rgb
    return (50 <= r <= 95) and (110 <= g <= 150) and (160 <= b <= 210)

def is_color_changed(base_rgb, current_rgb):
    """比對當前顏色與點擊前是否有顯著差異"""
    return any(abs(b - c) > 15 for b, c in zip(base_rgb, current_rgb))

def log_to_excel(file_name, cycle_num, all_records):
    """檢查並寫入數據至 Excel 檔案的下一列空格"""
    headers = ['輪次', '按鈕編號', '前鈕恢復時間(ms)', '新鈕變色時間(ms)', '總耗時(ms)']
    
    # 判斷檔案是否存在
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "測試結果"
        ws.append(headers)  # 寫入標題列
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    # 將這一輪(1~8顆按鈕)最新一次的數據接續寫入
    for num in range(1, 9):
        totals = all_records[num]["total"]
        p_recovers = all_records[num]["prev_recover"]
        c_changes = all_records[num]["curr_change"]
        
        if len(totals) >= cycle_num:  # 確保該輪有成功紀錄到數據
            row_data = [
                f"第 {cycle_num} 輪",
                f"Btn {num}",
                round(p_recovers[cycle_num - 1], 2),
                round(c_changes[cycle_num - 1], 2),
                round(totals[cycle_num - 1], 2)
            ]
            ws.append(row_data)  # openpyxl 的 append 會自動尋找下一列空格寫入
            
    wb.save(file_name)

def measure_color_change_time(total_cycles=5, excel_file="Test_Results.xlsx"):
    # target_title = "RF Switch Tool V3.0_build_2605180942"
    target_title = "RF Switch Tool V3.0_build_2605181435"
    
    try:
        app = Application(backend="win32").connect(title=target_title)
    except Exception:
        print(f"找不到『{target_title}』視窗，請確認程式已啟動。")
        return
        
    main_win = app.window(title=target_title)
    main_win.set_focus()
    time.sleep(0.5)
    
    buttons = {}
    hdc = win32gui.GetDC(0)
    
    print("正在初始化按鈕座標資訊...")
    for num in range(1, 9):
        auto_id = f"CheckBox28_{num}"
        try:
            btn_obj = main_win.child_window(auto_id=auto_id, control_type="System.Windows.Forms.CheckBox")
            rect = btn_obj.rectangle()
            cx = (rect.left + rect.right) // 2
            cy = (rect.top + rect.bottom) // 2
            buttons[num] = {"obj": btn_obj, "x": cx, "y": cy}
        except Exception as e:
            print(f"警告：無法定位到 {auto_id}，請檢查元件名稱是否正確。")
            win32gui.ReleaseDC(0, hdc)
            return

    all_records = {num: {"prev_recover": [], "curr_change": [], "total": []} for num in range(1, 9)}
    timeout_limit = 2.0
    
    print(f"\n開始執行 8 個按鈕輪流點選測試 (共設定測試 {total_cycles} 輪) ...")
    
    buttons[8]["obj"].click_input()
    win32api.SetCursorPos((0, 0))
    time.sleep(0.5)
    
    for cycle in range(1, total_cycles + 1):
        print(f"\n--- 第 {cycle:03d} / {total_cycles} 輪測試 ---")
        
        for current_num in range(1, 9):
            prev_num = 8 if current_num == 1 else current_num - 1
            
            curr_btn = buttons[current_num]
            prev_btn = buttons[prev_num]
            
            base_curr_color = get_pixel_color(hdc, curr_btn["x"], curr_btn["y"])
            base_prev_color = get_pixel_color(hdc, prev_btn["x"], prev_btn["y"])
            
            curr_btn["obj"].click_input()
            start_time = time.perf_counter()
            win32api.SetCursorPos((0, 0))
            
            prev_recover_time = None
            curr_change_time = None
            success = False
            
            while True:
                curr_color = get_pixel_color(hdc, curr_btn["x"], curr_btn["y"])
                prev_color = get_pixel_color(hdc, prev_btn["x"], prev_btn["y"])
                current_time = time.perf_counter()
                
                if prev_recover_time is None and is_color_changed(base_prev_color, prev_color):
                    prev_recover_time = current_time
                
                if curr_change_time is None and is_color_changed(base_curr_color, curr_color):
                    curr_change_time = current_time
                    success = True
                    break
                
                if (current_time - start_time) > timeout_limit:
                    break
            
            if success:
                if prev_recover_time is None:
                    prev_recover_time = current_time
                    
                p_recover_ms = (prev_recover_time - start_time) * 1000
                c_change_ms = (curr_change_time - start_time) * 1000
                total_ms = (curr_change_time - start_time) * 1000
                
                all_records[current_num]["prev_recover"].append(p_recover_ms)
                all_records[current_num]["curr_change"].append(c_change_ms)
                all_records[current_num]["total"].append(total_ms)
                
                print(f"  [按鈕 {current_num}] 前鈕回復: {p_recover_ms:.1f}ms | 新鈕變色: {c_change_ms:.1f}ms | 總計: {total_ms:.1f}ms")
            else:
                # 逾時也填入空數據維持格式一致
                all_records[current_num]["prev_recover"].append(0.0)
                all_records[current_num]["curr_change"].append(0.0)
                all_records[current_num]["total"].append(0.0)
                print(f"  [按鈕 {current_num}] 測試逾時，顏色未發生顯著改變。")
                
            time.sleep(0.05)
            
        # 【修改點】每一輪結束，立即將該輪數據存入 Excel 後一列
        log_to_excel(excel_file, cycle, all_records)
        print(f"  第 {cycle:03d} 輪測試數據已即時寫入 Excel。")

        print("\n" + "-"*85)
        print(f" 【 第 {cycle:03d} 輪結束即時累計統計結果 】")
        print("-"*85)
        print(f"{'按鈕編號':<10}{'成功次數':<10}{'前鈕恢復平均(ms)':<16}{'新鈕變色平均(ms)':<16}{'總耗時平均(ms)':<15}{'最快(ms)':<10}{'最慢(ms)':<10}")
        for num in range(1, 9):
            totals = [t for t in all_records[num]["total"] if t > 0] # 排除逾時的0
            p_recovers = [p for p in all_records[num]["prev_recover"] if p > 0]
            c_changes = [c for c in all_records[num]["curr_change"] if c > 0]
            if totals:
                print(f"Btn {num:<6}{len(totals):<12}{np.mean(p_recovers):<18.2f}{np.mean(c_changes):<18.2f}{np.mean(totals):<16.2f}{np.min(totals):<12.2f}{np.max(totals):<12.2f}")
            else:
                print(f"Btn {num:<6}{0:<12}{'N/A':<18}{'N/A':<18}{'N/A':<16}{'N/A':<12}{'N/A':<12}")
        print("="*85)
            
    win32gui.ReleaseDC(0, hdc)

if __name__ == "__main__":
    # 你可以在這裡更改儲存的 Excel 檔名
    measure_color_change_time(total_cycles=5, excel_file="RF_Switch_Test_Results.xlsx")