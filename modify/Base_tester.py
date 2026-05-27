import os
import time
import ctypes
import win32gui
import win32api
from pywinauto.application import Application
from openpyxl import load_workbook, Workbook

ctypes.windll.shcore.SetProcessDpiAwareness(2)

class RFTesterCore:
    def __init__(self, window_title):
        self.window_title = window_title
        self.app = None
        self.main_win = None
        self.hdc = None
        self.buttons = {}  # 存放映射後的按鈕物件與點位
    
    def connect_window(self):
        try:
            self.app = Application(backend="win32").connect(title=self.window_title)
            self.main_win = self.app.window(title=self.window_title)
            self.main_win.set_focus()
            self.hdc = win32gui.GetDC(0)
            print(f"成功連接視窗: {self.window_title}")
        except Exception as e:
            raise RuntimeError(f"無法連接到視窗 [{self.window_title}]: {e}")
        
    @staticmethod
    def _get_pixel_color(hdc, x, y):
        color = win32gui.GetPixel(hdc, x, y)
        return color & 0xff, (color >> 8) & 0xff, (color >> 16) & 0xff
    
    @staticmethod
    def _is_steel_blue(rgb):
        r, g, b = rgb
        return (r < 110) and (b > 150) and (b > r + 40)

    @staticmethod
    def _is_slate_gray(rgb):
        r, g, b = rgb
        return (80 < r < 150) and (abs(r - g) < 20) and (abs(r - b) < 20)
    
    def register_buttons(self, button_id_list, control_type="System.Windows.Forms.CheckBox"):
        if not self.main_win:
            self.connect_window()

        self.buttons = {}
        print("動態掃描並映射按鈕座標...")
        
        for index, auto_id in enumerate(button_id_list):
            try:
                btn = self.main_win.child_window(auto_id=auto_id, control_type=control_type)
                r = btn.rectangle()
                cx, cy = (r.left + r.right) // 2, (r.top + r.bottom) // 2
                
                self.buttons[index] = {
                    "id_name": auto_id,
                    "obj": btn,
                    "points": [(cx, cy-8), (cx, cy-12), (cx+10, cy-8), (cx-10, cy-8)]
                }
            except Exception as e:
                print(f"警告: 無法定位按鈕 {auto_id}, 錯誤: {e}")
    
    def measure_switch(self, curr_idx, prev_idx, timeout=2.0):
        """量測單次按鈕切換的反應時間"""
        curr_b = self.buttons[curr_idx]
        prev_b = self.buttons[prev_idx]

        win32api.SetCursorPos((0, 0))
        curr_b["obj"].click_input()
        t0 = time.perf_counter()
        win32api.SetCursorPos((0, 0))

        t_prev, t_curr = None, None
        while (time.perf_counter() - t0) < timeout:
            now = time.perf_counter()
            
            if t_curr is None:
                for px, py in curr_b["points"]:
                    if self._is_steel_blue(self._get_pixel_color(self.hdc, px, py)):
                        t_curr = now
                        break
            
            if t_prev is None:
                for px, py in prev_b["points"]:
                    if self._is_slate_gray(self._get_pixel_color(self.hdc, px, py)):
                        t_prev = now
                        break
            
            if t_curr:
                break

        res_prev = (t_prev - t0) * 1000 if t_prev else 0
        res_curr = (t_curr - t0) * 1000 if t_curr else 0
        return res_prev, res_curr

    def log_to_excel(self, file_name, cycle_num, record_row):
        headers = ['輪次', '邏輯順序', '按鈕實體ID', '前鈕恢復時間(ms)', '新鈕變色時間(ms)', '總耗時(ms)']
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "測試結果"
            ws.append(headers)
        else:
            wb = load_workbook(file_name)
            ws = wb.active

        ws.append(record_row)
        wb.save(file_name)

    def close(self):
        if self.hdc:
            win32gui.ReleaseDC(0, self.hdc)